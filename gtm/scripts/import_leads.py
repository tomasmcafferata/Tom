#!/usr/bin/env python3
"""
LEADS IMPORTER — Generic engine
================================
Reads a source Google Sheet (or xlsx in Drive) and imports leads into a
destination CRM sheet. The column mapping is passed via a JSON config file.

Usage:
    py -3 gtm/scripts/import_leads.py \
        --source-id  <sheet_id> \
        --dest-id    <sheet_id> \
        --dest-tab   Leads \
        --batch      "Ops ICP Ploteo" \
        --prefix     NDC \
        --mapping    gtm/scripts/formats/linkedin_sales_nav.json
        [--xlsx]     # add this flag if the source is an xlsx file in Drive

Mapping JSON supports:
    column_map       { dest_col: source_col }   direct field maps
    defaults         { dest_col: value }         static values for every row
    computed_fields  { dest_col: spec }          derived fields (see below)

Computed field specs:
    { "type": "split_first", "source": "Contacto" }   first word of field
    { "type": "split_rest",  "source": "Contacto" }   everything after first word
    { "type": "concat", "sources": ["A","B"], "separator": " | " }
"""

import argparse
import json
import sys
import io
import gspread
from google.oauth2.service_account import Credentials
from datetime import date

CREDENTIALS_PATH = r"C:\Users\Tomas Cafferata\Tom\credentials\service_account.json"

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]


# ─── CONNECTION ──────────────────────────────────────────────────────────────

def connect():
    creds = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=SCOPES)
    gc = gspread.authorize(creds)
    return gc, creds


# ─── READ SOURCE (Google Sheet) ──────────────────────────────────────────────

def read_source_gsheet(gc, source_id: str, source_tab: str = None) -> list[dict]:
    print(f"  Reading Google Sheet...")
    ss = gc.open_by_key(source_id)
    if source_tab:
        ws = ss.worksheet(source_tab)
        print(f"  Tab: '{ws.title}'")
    else:
        ws = ss.get_worksheet(0)
        print(f"  Tab: '{ws.title}' (first tab)")
    rows = [r for r in ws.get_all_records() if any(v for v in r.values())]
    print(f"  {len(rows)} rows found.")
    return rows


# ─── READ SOURCE (xlsx in Drive) ─────────────────────────────────────────────

def read_source_xlsx(creds, source_id: str) -> list[dict]:
    print(f"  Downloading xlsx from Drive...")
    try:
        import openpyxl
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaIoBaseDownload
    except ImportError as e:
        print(f"  [ERROR] Missing dependency: {e}")
        print(f"  Run: py -3 -m pip install openpyxl google-api-python-client")
        sys.exit(1)

    service = build("drive", "v3", credentials=creds)
    request = service.files().get_media(fileId=source_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()

    fh.seek(0)
    wb = openpyxl.load_workbook(fh)
    ws = wb.active
    headers = [str(cell.value) if cell.value is not None else "" for cell in ws[1]]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v for v in row):
            rows.append({h: (str(v) if v is not None else "") for h, v in zip(headers, row)})

    print(f"  {len(rows)} rows found.")
    return rows


# ─── DESTINATION SCHEMA ──────────────────────────────────────────────────────

def get_dest_worksheet(gc, dest_id: str, dest_tab: str):
    sh = gc.open_by_key(dest_id)
    try:
        ws = sh.worksheet(dest_tab)
    except gspread.exceptions.WorksheetNotFound:
        print(f"  [ERROR] Tab '{dest_tab}' not found.")
        print(f"  Available: {[w.title for w in sh.worksheets()]}")
        sys.exit(1)
    headers = ws.row_values(1)
    print(f"  Destination has {len(headers)} columns.")
    return ws, headers


def get_next_lead_id(ws, prefix: str) -> int:
    existing = ws.col_values(1)
    matches = [v for v in existing if v.startswith(f"{prefix}-")]
    if not matches:
        return 1
    nums = []
    for v in matches:
        try:
            nums.append(int(v.split("-")[1]))
        except (IndexError, ValueError):
            pass
    return max(nums) + 1 if nums else 1


# ─── TRANSFORM ───────────────────────────────────────────────────────────────

def apply_computed(row: dict, computed_fields: dict) -> dict:
    """Resolve computed fields from source row."""
    result = {}
    for dest_col, spec in computed_fields.items():
        t = spec.get("type")

        if t == "split_first":
            val = str(row.get(spec["source"], "") or "").strip()
            result[dest_col] = val.split()[0] if val else ""

        elif t == "split_rest":
            val = str(row.get(spec["source"], "") or "").strip()
            parts = val.split()
            result[dest_col] = " ".join(parts[1:]) if len(parts) > 1 else ""

        elif t == "concat":
            parts = [str(row.get(s, "") or "").strip() for s in spec.get("sources", [])]
            parts = [p for p in parts if p]
            result[dest_col] = spec.get("separator", " | ").join(parts)

        elif t == "map_value":
            val = str(row.get(spec["source"], "") or "").strip()
            lookup = spec.get("map", {})
            # Match by checking if the source value starts with any key
            matched = spec.get("default", "")
            for key, mapped in lookup.items():
                if val.startswith(key):
                    matched = mapped
                    break
            result[dest_col] = matched

    return result


def clean_email(value: str) -> str:
    """If a cell has multiple emails separated by / or , take the first."""
    for sep in [" / ", "/", ","]:
        if sep in value:
            return value.split(sep)[0].strip()
    return value.strip()


def transform(
    source_rows: list[dict],
    dest_headers: list[str],
    column_map: dict,
    defaults: dict,
    computed_fields: dict,
    prefix: str,
    batch: str,
    start_id: int,
) -> list[list]:
    today = date.today().strftime("%Y-%m-%d")
    auto = {"source_batch": batch, "import_date": today}

    output = []
    for i, row in enumerate(source_rows):
        lead_id = f"{prefix}-{str(start_id + i).zfill(3)}" if prefix else ""
        computed = apply_computed(row, computed_fields)

        built = {}
        for col in dest_headers:
            if col == "lead_id" and prefix:
                built[col] = lead_id
            elif col in computed:
                built[col] = computed[col]
            elif col in column_map:
                val = str(row.get(column_map[col], "") or "")
                # Clean emails
                if col == "email":
                    val = clean_email(val)
                built[col] = val
            elif col in defaults:
                built[col] = defaults[col]
            elif col in auto:
                built[col] = auto[col]
            else:
                built[col] = ""

        output.append([built[col] for col in dest_headers])

    return output


# ─── WRITE ───────────────────────────────────────────────────────────────────

def write_rows(ws, rows: list[list]):
    print(f"  Writing {len(rows)} rows...")
    ws.append_rows(rows, value_input_option="USER_ENTERED")
    print(f"  Done.")


# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generic leads importer")
    parser.add_argument("--source-id",  required=True)
    parser.add_argument("--dest-id",    required=True)
    parser.add_argument("--dest-tab",   default="Leads")
    parser.add_argument("--batch",      required=True)
    parser.add_argument("--prefix",     required=True)
    parser.add_argument("--mapping",    required=True)
    parser.add_argument("--xlsx",       action="store_true", help="Source is an xlsx file in Drive")
    parser.add_argument("--source-tab", default=None, help="Tab name in source sheet (default: first tab)")
    args = parser.parse_args()

    with open(args.mapping) as f:
        mapping = json.load(f)
    column_map      = mapping.get("column_map", {})
    defaults        = mapping.get("defaults", {})
    computed_fields = mapping.get("computed_fields", {})

    print("\n" + "="*55)
    print("  LEADS IMPORTER")
    print("="*55)

    gc, creds = connect()
    print("\n[1/4] Connected.")

    print("\n[2/4] Reading source...")
    if args.xlsx:
        source_rows = read_source_xlsx(creds, args.source_id)
    else:
        source_rows = read_source_gsheet(gc, args.source_id, args.source_tab)

    print("\n[3/4] Reading destination...")
    dest_ws, dest_headers = get_dest_worksheet(gc, args.dest_id, args.dest_tab)
    start_id = get_next_lead_id(dest_ws, args.prefix)
    print(f"  Next ID: {args.prefix}-{str(start_id).zfill(3)}")

    print("\n[4/4] Transforming and writing...")
    rows = transform(source_rows, dest_headers, column_map, defaults, computed_fields, args.prefix, args.batch, start_id)
    write_rows(dest_ws, rows)

    end_id = start_id + len(rows) - 1
    print(f"\n  OK: {len(rows)} leads imported.")
    print(f"  IDs: {args.prefix}-{str(start_id).zfill(3)} to {args.prefix}-{str(end_id).zfill(3)}")
    print("="*55 + "\n")


if __name__ == "__main__":
    main()
